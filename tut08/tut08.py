# -*- coding: utf-8 -*-
"""Assignment_lab.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1f53I5VmFMLyGDrdzgdJmRCp9ENCBXYd-
"""

pip install pandas openpyxl

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#PATHS
attendance_path="input_attendance.csv"
name_path="stud_list.txt"

def find(a,i):
  if len(a)==0:
    return 0
  cnt=0
  for j in a:
    if j==i:
      cnt+=1
  return cnt

classes_taken_dates = ["06-08-2024", "13-08-2024", "20-08-2024", "27-08-2024", "03-09-2024", "17-09-2024", "01-10-2024"]

xindex=dict() #index in xcell file

df = pd.read_csv("input_attendance.csv")
df['Timestamp'] = pd.to_datetime(df['Timestamp'], format='%d/%m/%Y %H:%M:%S')
df['Time'] = df['Timestamp'].dt.time
start_time = pd.to_datetime("18:00:00").time()
end_time = pd.to_datetime("20:00:00").time()
mask = (df['Time'] >= start_time) & (df['Time'] <= end_time)
filtered_df = df.loc[mask]

with open("stud_list.txt", 'r') as file:
    content = file.read()

name=dict()
lines = content.strip().split('\n')
for line in lines:
    parts = line.split(' ', 1)  # Split on the first space
    if len(parts) == 2:
        roll = parts[0]  # First part is the roll number
        name1 = parts[1]  # Keep the entire name including spaces
        name[roll] = name1

present=dict() #which are the dates
tot=dict() #total present
for i in name:
  present[i]=[]
  tot[i]=0
for i in range(len(filtered_df)):
    roll_value = filtered_df['Roll'].iloc[i]

    if isinstance(roll_value, str) and len(roll_value) >= 8:
        roll = roll_value[:8]
    else:
        continue
    tot[roll] += 1

    day = filtered_df['Timestamp'].dt.strftime('%d-%m-%Y').iloc[i]
    if day in classes_taken_dates:
        present[roll].append(day)

dataf={
    'Roll No':[],
    'Name':[],
}
for i in classes_taken_dates:

  dataf[i]=[]

dataf['Total attendance in all classes']=[];
dataf['Total attendance marked']=[];
dataf['Total attendance allowed']=[];
dataf['Proxy']=[];
for i in name:
  dataf['Roll No'].append(i)
  dataf['Name'].append(name[i])
  for j in classes_taken_dates:
    dataf[j].append(find(present[i],j))
  dataf['Total attendance marked'].append(tot[i])
  dataf['Total attendance allowed'].append(len(classes_taken_dates)*2)
  dataf['Proxy'].append(0)
  dataf['Total attendance in all classes'].append(0);
# dataf
xcell_data=pd.DataFrame(dataf)
excel_file = 'Attendance_output.xlsx'
xcell_data.to_excel(excel_file, index=False,sheet_name='attendance')

wb = load_workbook(excel_file)
ws = wb['attendance']

red_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')  # Soft Red
green_fill = PatternFill(start_color='66FF66', end_color='66FF66', fill_type='solid')  # Soft Green
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Soft Yellow

def fill(a,row):
   val=0
   if(ws[a+f'{row}'].value):
    val=ws[a+f'{row}'].value
   if(val > 2):
    ws[a+f'{row}'].fill = red_fill
   elif(val==2):
    ws[a+f'{row}'].fill = green_fill
   elif(val==1):
    ws[a+f'{row}'].fill = yellow_fill

for row in range(2, len(df) + 2):
    # starting from row 2 to account for the header
    ws[f'J{row}'] = f'=C{row}+D{row}+E{row}+F{row}+G{row}+H{row}+I{row}'
    ws[f'M{row}'] = f'=K{row}-J{row}'

    fill('C',row)
    fill('D',row)
    fill('E',row)
    fill('F',row)
    fill('G',row)
    fill('H',row)
    fill('I',row)


wb.save(excel_file)